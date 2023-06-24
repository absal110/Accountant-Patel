from openpyxl import load_workbook
import pyautogui
import time

wb = load_workbook(r'D:\Data\SHARES.xlsx')
ws = wb.active
# max = ws.max_row

i = 7
sharesSaved = []
while ws[f"K{i}"].value!=None:
    shareName = ws[f"K{i}"].value.split("-")
    sharesSaved.append(shareName[0].strip())
    i+=1
# print(sharesSaved)

i=7
sharesToBeSaved = []
while ws[f"B{i}"].value!=None:
    if ws[f"B{i}"].value not in sharesSaved:
        sharesToBeSaved.append(ws[f"B{i}"].value)
    i+=1

remove = []
futuresandoptions = []
for al_each in sharesToBeSaved:
    each = al_each.split()
    if len(each)>2:
        # print(al_each)
        if each[0]=="Fut":
            if "F "+each[1] not in futuresandoptions and "F "+each[1] not in sharesSaved:
                futuresandoptions.append("F "+each[1])
            remove.append(al_each)
        elif each[0]=="Opt":
            if "O "+each[1] not in futuresandoptions and "O "+each[1] not in sharesSaved:
                futuresandoptions.append("O "+each[1])
            remove.append(al_each)
for each in remove:
    sharesToBeSaved.remove(each)
print(sharesToBeSaved)
print(futuresandoptions)

xyz = input("Writing share names press enter")

time.sleep(5)
for each in futuresandoptions:
# for i in range(1):
#     each = "MuzammilPatel patel raju"
    pyautogui.typewrite(each+"-BP EQ")
    pyautogui.press('enter')
    # pyautogui.press('enter')
    pyautogui.press('enter')
    pyautogui.typewrite("F")
    pyautogui.press('enter')
    pyautogui.typewrite("NO")
    pyautogui.press('enter')
    pyautogui.press('enter')
    pyautogui.press('enter')
    time.sleep(0.5)
    pyautogui.press('enter')

