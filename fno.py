from openpyxl import load_workbook
import pyautogui
import time

fName = "01.06.2022"

wb = load_workbook(rf'D:\SSTallyentrymaterial\fno\{fName}.xlsx')
ws = wb.active
max = ws.max_row

i = 7
buy = [[]]
sell = [[]]
b=[0,0]
s=[0,0]
stt = 0

while i<= max:
    if ws[f'B{i}'].value != None:
        if ws[f'D{i}'].value != 0:
            b[0] += ws[f'D{i}'].value
            if ws[f'H{i}'].value<0:
                b[1] += ws[f'H{i}'].value*-1
            else:
                b[1] += ws[f'H{i}'].value
        elif ws[f'E{i}'].value != 0:
            s[0] += ws[f'E{i}'].value
            if ws[f'H{i}'].value < 0:
                s[1] += ws[f'H{i}'].value * -1
            else:
                s[1] += ws[f'H{i}'].value
        else:
            while ws[f'H{i}'].value != None:
                stt +=ws[f'H{i}'].value
                i+=1
            break
        if b[0]>0:
            buy.append([ws[f'B{i}'].value,b[0],b[1]])
        if s[0]>0:
            sell.append([ws[f'B{i}'].value,s[0],s[1]])
        b = [0, 0]
        s = [0, 0]
    i+=1


buy.pop(0)
sell.pop(0)

print(buy)
print(sell)

x='N'

if buy:
    for each in buy:
        a = each[0].split()
        each[0] = a[:1]
else:
    while 1:
        x = input("No purchase plese press Y")
        if x == 'Y':
            break

for each in sell:
    a = each[0].split()
    each[0] = a[:2]

shares = []
for each in buy:
    if each[0] not in shares:
        shares.append(each[0])
for each in sell:
    if each[0] not in shares:
        shares.append(each[0])


dueToUs = ws[f'H{max}'].value
if dueToUs<0:
    dueToUs=dueToUs*-1
filekaname = ws["A3"].value
billDate = filekaname[-10:]

print(shares)

# print(billDate)
print(dueToUs,stt)

suffix = input("Enter Suffix if any or press 'ENTER'\n")

time.sleep(2.5)


pyautogui.keyDown('alt')
time.sleep(0.25)
pyautogui.press('tab')
pyautogui.keyUp('alt')
time.sleep(0.65)
pyautogui.keyDown('alt')
pyautogui.press('a')
pyautogui.keyUp('alt')
pyautogui.press('f2')
pyautogui.typewrite(billDate)
pyautogui.press('enter')

if x!='Y':
    pyautogui.typewrite("SHARE PURCHASE")
    pyautogui.press('enter')
    for each in buy:
        pyautogui.typewrite(each[0])
        if suffix!="":
            pyautogui.typewrite(suffix)
        pyautogui.press('enter')
        pyautogui.typewrite(str(each[1]))
        pyautogui.press('enter')
        pyautogui.press('enter')
        pyautogui.press('enter')
        pyautogui.typewrite(str(each[2]))
        pyautogui.press('enter')
    pyautogui.press('enter')
    pyautogui.typewrite("To")
    pyautogui.press('enter')

if sell:
    pyautogui.typewrite("SHARE SALE")
    pyautogui.press('enter')

    for each in sell:
        pyautogui.typewrite(each[0])
        if suffix != "":
            pyautogui.typewrite(suffix)
        pyautogui.press('enter')
        pyautogui.typewrite(str(each[1]))
        pyautogui.press('enter')
        pyautogui.press('enter')
        pyautogui.press('enter')
        pyautogui.typewrite(str(each[2]))
        pyautogui.press('enter')
    pyautogui.press('enter')
    pyautogui.press('enter')
    pyautogui.typewrite("BP")
    pyautogui.press('enter')
    pyautogui.typewrite(str(dueToUs))
    pyautogui.press('enter')
    pyautogui.press('enter')
    pyautogui.typewrite("STT")
    pyautogui.press('enter')
    pyautogui.typewrite(str(stt))
    pyautogui.press('enter')
    pyautogui.press('enter')
    pyautogui.typewrite("R")
    pyautogui.press('enter')
    pyautogui.press('enter')
    pyautogui.typewrite(fName)
    pyautogui.press('enter')
    time.sleep(0.75)
    pyautogui.press('enter')